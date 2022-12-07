# -*- coding: utf-8 -*-
"""
Created on Thu Dec 17 17:22:10 2020

@author: pdsu_szj
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Dec 16 15:32:21 2020
read all files
@author: pdsu_szj
"""

import networkx as nx
from sklearn.metrics.cluster import normalized_mutual_info_score
from sklearn.metrics.cluster import adjusted_rand_score
from collections import defaultdict
import time
import datetime
from sklearn import metrics
import math
import matplotlib.pyplot as plt  # 画图用
import numpy as np
from numpy import linalg as LA
import openpyxl
#global comm_list
#comm_list=[]


def str_to_int(x):
    return [[int(v) for v in line.split()] for line in x]
def node_addition(G,addnodes,communitys):	#输入的communitys社区格式为｛节点：社区名称｝
    change_comm=set()#存放结构可能发现改变的社区标签
    processed_edges=set()#已处理过的边，需要从增加边中删除已处理过的边
    
    for u in addnodes:
        neighbors_u=G.neighbors(u)
        neig_comm=set()#邻居所在社区标签
        pc=set()       
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            pc.add((u,v))
            pc.add((v,u))#无向图中都 是一条边，加两次方便操作
        if len(neig_comm)>1:   #说明此加入结点不在社区内部
           change_comm=change_comm | neig_comm
           lab=max(communitys.values())+1
           communitys.setdefault(u,lab)#为u分配一个社区标签
           change_comm.add(lab)
        else:
            if len(neig_comm)==1:#说明结点在社区内部，或只与一个社区连接
               communitys.setdefault(u,list(neig_comm)[0])#将结点加入到本社区
               processed_edges=processed_edges|pc
            else:
               communitys.setdefault(u,max(communitys.values())+1)#新加结点未和其它结点有连接，分配新的社区标签
    
    return change_comm,processed_edges,communitys# 返回可能发生变化的社区，处理过的边和最新社区结构。
def node_deletion(G,delnodes,communitys):					#tested, correct
    change_comm=set()#存放结构可能发现改变的社区标签
    processed_edges=set()#已处理过的边，需要从删除边中删除已处理过的边
    for u in delnodes:
        neighbors_u=G.neighbors(u)
        neig_comm=set()#邻居所在社区标签        
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            processed_edges.add((u,v))
            processed_edges.add((v,u))
        del communitys[u] #删除结点和社区
        change_comm=change_comm | neig_comm       
    return change_comm,processed_edges,communitys# 返回可能发生变化的社区，处理过的边和最新社区结构。

def edge_addition(addedges,communitys):#如果加入边在社区内部，不会引起社区变化则不做处理，否则标记
    change_comm=set()#存放结构可能发现改变的社区标签
#    print addedges
#    print communitys    
    for item in addedges:        
        neig_comm=set()#邻居所在社区标签 
        neig_comm.add(communitys[item[0]])#判断一边两端的节点所在社区
        neig_comm.add(communitys[item[1]])
        if len(neig_comm)>1:   #说明此加入边不在社区内部
           change_comm=change_comm | neig_comm
    return change_comm # 返回可能发生变化的社区，

def edge_deletion(deledges,communitys):#如果删除边在社区内部可能引起社区变化，在社区外部则不会变化
    change_comm=set()#存放结构可能发现改变的社区标签    
    for item in deledges:        
        neig_comm=set()#邻居所在社区标签 
        neig_comm.add(communitys[item[0]])#判断一边两端的节点所在社区
        neig_comm.add(communitys[item[1]])
        if len(neig_comm)==1:   #说明此加入边不在社区内部
           change_comm=change_comm | neig_comm

    return change_comm # 返回可能发生变化的社区

def getchangegraph(all_change_comm,newcomm,Gt):
    Gte=nx.Graph()
    com_key=newcomm.keys()
    for v in Gt.nodes():
        if v not in com_key or newcomm[v] in all_change_comm:            
            Gte.add_node(v)
            neig_v= Gt.neighbors(v)        
            for u in neig_v:                          
               if u not in com_key or newcomm[u] in all_change_comm:                   
                   Gte.add_edge(v,u)
                   Gte.add_node(u)          
    
    return Gte
#cdme社区检测****************************************************************

nodecount_comm = defaultdict(int) #全局变量
def CDME(G):
    deg = G.degree()
    def AA(NA,NB):
        comm_nodes=list(NA & NB)
        sim=0
        for node in comm_nodes:
            degnode= deg[node]
            if  deg[node]==1:
               degnode=1.1              
            sim=sim+(1.0/math.log(degnode))
        return sim    
    #Compute the jaccard similarity coefficient of two node
    def simjkd( u, v):        
        set_v = set( G.neighbors(v))
        set_v.add(v)
        set_u = set( G.neighbors(u))
        set_u.add(u)
        jac = len(set_v & set_u) * 1.0 / len(set_v | set_u)    
        return jac
    # Initialize communities of each node
    #第一阶段，一个节点为一个社区
    node_community = dict(zip(G.nodes(), G.nodes()))
    #Compute the core groups 
    #第二个阶段，计算核心分组
    st = {}  # storge the AA
    #compute the AA
    for node in G.nodes():
        Nv=sorted(G.neighbors(node))
        for u in Nv:
            Nu=G.neighbors(u)
            keys = str(node) + '_' + str(u)
            st.setdefault(keys,AA(set(Nv), set(Nu)))
    print('AAindex,done')  
    for node in G.nodes():
        #The degree of each node
        deg_node = deg[node]
        flag = True
        maxsimdeg = 0            
        selected = node            
        if deg_node == 1:
            # node_community[node] =  node_community[ G.neighbors(node)[0]]
             node_community[node] =  node_community[list( G.neighbors(node))[0]]
        else:               
            for neig in  G.neighbors(node):                   
                deg_neig =  deg[neig]                    
                if flag is True and deg_node <= deg_neig:
                    flag = False
                    break
      
            if flag is False:
                for neig in sorted( G.neighbors(node)):                
                    deg_neig =  deg[neig]
                    # Compute the Jaccard similarity coefficient
                    #nodesim =  simjkd(node, neig)
                    #Use the AAindex
                    keys = str(node) + '_' + str(neig)
                    nodesim= st[keys]  
                    # Compute the node attraction
                    nodesimdeg = deg_neig*nodesim
                    if nodesimdeg > maxsimdeg:
                        selected = neig
                        maxsimdeg = nodesimdeg
                    node_community[node] =  node_community[selected]  
    old_persum = -(2 ** 63 - 1)
    old_netw_per = -(2 ** 63 - 1)       

    persum = old_persum + 1
    netw_per = old_netw_per + 0.1
    maxit = 5
    itern = 0
    
    print("loop begin:")        
    while itern < maxit:            
        itern += 1
        old_netw_per = netw_per
        old_persum = persum
        persum = 0            
        for node in G.nodes():
            neiglist =sorted( G.neighbors(node))
            cur_p = per(G,node,node_community)#               
            nodeneig_comm = nodecount_comm.keys()                
            cur_p_neig = 0
    
            for neig in neiglist:
                cur_p_neig += per(G,neig,node_community)
    #                #计算当前节点与邻居节点构成三角形数量
    #                for neig in neiglist:
    #                    tri=set(neiglist)&set(self.G.neighbors(neig))
    #                    cur_tri += tri               
    #                #度加2倍三角形
    #                cur_p_neig+=2*cur_tri
            for neig_comm in nodeneig_comm:
               
                node_pre_comm = node_community[node]                   
                new_p_neig = 0                    
                if node_pre_comm != neig_comm:
                    node_community[node] = neig_comm
                    new_p = per(G,node,node_community)
                    
                    if cur_p <= new_p:
    
                        if cur_p == new_p:
                            for newneig in neiglist:
                                new_p_neig += per(G,newneig,node_community)                                    
                            if cur_p_neig < new_p_neig:
                                cur_p = new_p
                                cur_p_neig = new_p_neig
                            else:
                                node_community[node] = node_pre_comm
    
                        else:
                            for newneig in neiglist:
                                new_p_neig += per(G,newneig,node_community)
                            cur_p = new_p
                            cur_p_neig = new_p_neig
                    else:
                        node_community[node] = node_pre_comm    
            persum += cur_p
#            print(node_community)
    print ("loop done")
    #转换社区形式，{结点：社区}转换为  {社区：[结点]} 
#    graph_result = defaultdict(list)
#    for item in node_community.keys():
#        node_comm = int(node_community[item])
#        graph_result[node_comm].append(item)
    


    return node_community
# The internal degree of node v in a community
def per(G,v,node_community):
    
    neiglist1 = G.neighbors(v)     
    in_v = 0
    #tri=0
    global nodecount_comm
   
    for neig in neiglist1:
        if node_community[neig] == node_community[v]:
            in_v += 1
            #tri+=len(set(neiglist1)&set(self.G.neighbors(neig)))
        else:
            
            nodecount_comm[node_community[neig]] += 1
    
    cin_v = 1.0*(in_v*in_v)
    per = cin_v
    return per
    
#****************************************************************************

def Errorrate(clusters, classes,n):
    #计算错误率，公式||A*A'-C*C'||,A'代表矩阵的转置。A代表发现的社区，行为节点，列为社区，结点属于一个社区，则对应交叉单元格为1，否则为0，C为真实社区的矩阵
    A=np.zeros((n,len(clusters)),int)
    C=np.zeros((n,len(classes)),int)
    k=0
    for nodelist in clusters:
        for node in nodelist:
            A[node-1][k]=1
        k=k+1
    k=0
    for nodelist in classes:
        for node in nodelist:
            C[node-1][k]=1
        k=k+1
    t=A.dot(A.T)-C.dot(C.T)
    errors=LA.norm(t)
    return errors
    
def purity_score(clusters, classes):
    """
    Calculate the purity score for the given cluster assignments and ground truth classes
    
    :param clusters: the cluster assignments array
    :type clusters: numpy.array
    
    :param classes: the ground truth classes
    :type classes: numpy.array
    
    :returns: the purity score
    :rtype: float
    """
    
    A = np.c_[(clusters,classes)]

    n_accurate = 0.

    for j in np.unique(A[:,0]):
        z = A[A[:,0] == j, 1]
        x = np.argmax(np.bincount(z))
        n_accurate += len(z[z == x])

    return n_accurate / A.shape[0]
def conver_comm_to_lab(comm1):#转换社区格式为，标签为主键，节点为value
    overl_community={}
    for node_v,com_lab in comm1.items():
        if com_lab in overl_community.keys():
            overl_community[com_lab].append(node_v)
        else:
            overl_community.update({com_lab:[node_v]})
    return overl_community
def getscore(comm_true,comm_dete,num):
    actual=[]
    baseline=[]
    for j in range(len(comm_true)):#groundtruth，j代表每个社区,j为社区名称
    	for c in comm_true[j]: #社区中的每个节点，代表各节点
    		flag=False
    		for k in range(len(comm_dete)): #检测到的社区，k为社区名称
    			if c in comm_dete[k] and flag==False:
    				flag=True 
    				actual.append(j)
    				baseline.append(k)
    				break
    
    NMI1= metrics.normalized_mutual_info_score(actual, baseline)
    ARI1=metrics.adjusted_rand_score(actual,baseline)
    Purity1=purity_score(baseline,actual)
    #errors=Errorrate(comm_dete,comm_true,num)
    errors=0
    print ('nmi',NMI1)
    print ('ari', ARI1)
    print('purity',Purity1)
    print('rate error',errors)
    
    return NMI1,ARI1,Purity1,errors
def drawcommunity(g,partition,filepath):    
    pos = nx.spring_layout(g)
    count1 = 0
    t=0
    node_color=['#66CCCC','#FFCC00','#99CC33','#CC6600','#CCCC66','#FF99CC','#66FFFF','#66CC66','#CCFFFF','#CCCC00','#CC99CC','#FFFFCC']
#    print(node_color[1])   
    
    for com in set(partition.values()) :
        count1 = count1 + 1.
        list_nodes = [nodes for nodes in partition.keys()
                                    if partition[nodes] == com]
    
        nx.draw_networkx_nodes(g, pos, list_nodes, node_size = 220,
                                    node_color = node_color[t])
        nx.draw_networkx_labels(g, pos)
        t=t+1
    
    nx.draw_networkx_edges(g,pos,with_labels = True, alpha=0.5 )
    plt.savefig(filepath)
    plt.show()



''''############################################################
#----------main-----------------
'''
edges_added = set()
edges_removed = set()
nodes_added = set()
nodes_removed = set()
G=nx.Graph()
allpath='./data/my_LFR/files.txt'
with open(allpath,'r') as f:
    allpathlist=f.readlines()
    f.close()
#allpathlists=allpathlist[0].strip('\n')
pathfile=''
for pt in allpathlist:
    pathfile=pt.strip('\n')
    print(pathfile)
    path='./data/my_LFR/'+pathfile+'/'
    edge_file=''
    comm_file=''
    G.clear()
    #read edgefile list, where storage the filename of each snapshot
    edgefilelist=[]
    commfilelist=[]
    with open(path+'edgeslist.txt','r') as f:
        edgefilelist=f.readlines()
        f.close()
    edge_file=edgefilelist[0].strip('\n')
    with open(path+'commlist.txt','r') as f:
        commfilelist=f.readlines()
        f.close()
    comm_file=commfilelist[0].strip('\n')
    
    #path='./LFR/t/'
    #path='./data/test/'
    with open(path+edge_file,'r') as f:
        edge_list=f.readlines()
        for edge in edge_list:
            edge=edge.split()
            G.add_node(int(edge[0]))
            G.add_node(int(edge[1]))
            G.add_edge(int(edge[0]),int(edge[1]))
        f.close()
    G=G.to_undirected()
    ##初始图
    #print('T0时间片的网络G0*********************************************')
    #nx.draw_networkx(G)
    #fpath='./data/pic/G_0.png'
    #plt.savefig(fpath)           #输出方式1: 将图像存为一个png格式的图片文件
    #plt.show()
    #print G.edges()
    #comm_file='switch.t01.comm'
    nodenumber=G.number_of_nodes()
    with open(path+comm_file,'r') as f:
        comm_list=f.readlines()
        comm_list=str_to_int(comm_list)#真实社区
        f.close()

    #第三个阶段，模拟马太效应阶段
    
    comm={}#用来存放所检测到的社区结构，格式｛节点：社区标签｝
    comm=CDME(G)   #初始社区 
    #画社区
    #print('T0时间片的社区C0*********************************************')
    #drawcommunity(G,comm,'./data/pic/community_0.png')
    initcomm=conver_comm_to_lab(comm)
    comm_va=list(initcomm.values())
    commu_num=len(comm_va)
    tru_num=len(comm_list)
    NMI,ARI,Purity,Errors=getscore(comm_list,comm_va,nodenumber)
    path_score='result_score_LFR.xlsx'
    #f = open(path_score,'a+')        #写入文件   
    #f.write('path'+"\t"+'NMI'+"\t"+'ARI'+"\t"+'Purity'+'\t'+'detected_community_number'+'ture_community_number'+'errors'"\n" )
    #f.write(path+'_1'+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
    #f.close()
    #wb=openpyxl.Workbook(path_score)
    #wb.save(path_score)
    wb=openpyxl.load_workbook(filename = path_score)
    ws=wb.create_sheet(path[14:len(path)-1])
    row=['path','NMI','ARI','Purity','detected_community_number','ture_community_number','errors']
    ws.append(row)
    row=['1',str(NMI),str(ARI),str(Purity),str(commu_num),str(tru_num),str(Errors)]
    ws.append(row)
    
    start=time.time()
    G1=nx.Graph()
    G2=nx.Graph()
    G1=G
    
    l=len(edgefilelist)
    for i in range(1,l):
        print('begin loop:', i)
        comm_new_file=open(path+commfilelist[i].strip('\n'),'r')
        comm_new=comm_new_file.readlines()
        comm_new_file.close()
        comm_new=str_to_int(comm_new)
    
        edge_list_new_file=open(path+edgefilelist[i].strip('\n'),'r')
        edge_list_new=edge_list_new_file.readlines()   
        edge_list_new_file.close()
    
        
    #    for line in edge_list_old:
    #         temp = line.strip().split()
    #      
    #         G1.add_edge(int(temp[0]),int(temp[1]))
        for line in edge_list_new:
             temp = line.strip().split()     
             G2.add_edge(int(temp[0]),int(temp[1]))
    
        #当前时间片和上一时间片总节点数，两集合相关
        total_nodes =set(G1.nodes()) | set(G2.nodes())
    
        nodes_added=set(G2.nodes())-set(G1.nodes())
        #print ('增加节点集为：',nodes_added)
        nodes_removed=set(G1.nodes())-set(G2.nodes())
        #print ('删除节点集为：',nodes_removed)
    
        edges_added = set(G2.edges())-set(G1.edges())
        #print ('增加边集为：',edges_added)
        edges_removed = set(G1.edges())-set(G2.edges())
        #print ('删除边集为：',edges_removed)
    
        all_change_comm=set()
        #添加结点处理#############################################################
        addn_ch_comm,addn_pro_edges,addn_commu = node_addition(G2,nodes_added,comm)
    
        edges_added=edges_added-addn_pro_edges#去掉已处理的边
    
        all_change_comm=all_change_comm | addn_ch_comm
    
       
        #删除结点处理#############################################################
    
        deln_ch_comm,deln_pro_edges,deln_commu  = node_deletion(G1,nodes_removed,addn_commu)
        all_change_comm=all_change_comm | deln_ch_comm
        edges_removed=edges_removed-deln_pro_edges
    
        #添加边处理#############################################################
    #    print('edges_added',edges_added)
        adde_ch_comm= edge_addition(edges_added,deln_commu)
        all_change_comm=all_change_comm | adde_ch_comm
    #    print('all_change_comm',all_change_comm)
        #删除边处理#############################################################
        dele_ch_comm= edge_deletion(edges_removed,deln_commu)
        all_change_comm=all_change_comm | dele_ch_comm
    #    print('all_change_comm',all_change_comm)
        unchangecomm=()#未改变的社区标签
        newcomm={}#格式为｛节点：社区｝
        newcomm=deln_commu# 添加边和删除边，只是在现有节点上处理，不会新增节点，删除节点（前面已处理）
        unchangecomm=set(newcomm.values())-all_change_comm
        unchcommunity={ key:value for key,value in newcomm.items() if value in unchangecomm}#未改变的社区 ：标签和结点
        #找出变化社区所对应的子图，然后对子图运用马太效应动力学找出新的社区结构，加上未改变的社区结构，得到新的社区结构。
    #    print('change community:',all_change_comm)
        Gtemp=nx.Graph()
        Gtemp=getchangegraph(all_change_comm,newcomm,G2)
        
        unchagecom_maxlabe=0    
        if len(unchangecomm)>0:
            unchagecom_maxlabe=max(unchangecomm)
    #    print('subG',Gtemp.edges())
        if Gtemp.number_of_edges()<1:#社区未发生变化
            comm=newcomm
        else:           
            getnewcomm=CDME(Gtemp)
            
            
    #        print('newcomm',getnewcomm)
            #合并社区结构，未改的加上新获得的
            mergecomm={}#合并字黄格式为｛节点：社区｝
            mergecomm.update(unchcommunity)
            mergecomm.update(getnewcomm)
            #mergecomm=dict(**unchcommunity, **getnewcomm )
            comm=mergecomm #把当前获得的社区结构，作为下一次的社区输入
            detectcom=list(conver_comm_to_lab(comm).values())
            commu_num=len(detectcom)
            tru_num=len(comm_new)
    #    print detectcom
        nodenumber = G2.number_of_nodes()
    #        print('T'+str(i-1)+'时间片的网络社团结构C'+str(i-1)+'*********************************************')
    #        print(unchcommunity)
    #        print(newcomm)
    #        print(comm)
    #        drawcommunity(G2,comm,'./data/pic/community_'+str(i-1)+'.png')
    #    print ('getcommunity:',conver_comm_to_lab(comm))
        NMI,ARI,Purity,Errors=getscore(comm_new,detectcom,nodenumber)   #第一个参数据为真实社区，第二个为检测到的社区
    #    print('community number:', len(set(comm.values())))
    #    print(comm)
        G1.clear()
        G1.add_edges_from(G2.edges())
        G2.clear()
    #    f = open(path_score,'a+')        #写入文件   
    #    f.write(path+'_'+str(i)+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
    #    f.close()
        row=[str(i+1),str(NMI),str(ARI),str(Purity),str(commu_num),str(tru_num),str(Errors)]
        ws.append(row)
    wb.save(path_score)    
print ('all done')

#edge1=set()
#edge2=set()
#edge1.add((1,2))
#edge1.add((1,3))
#edge1.add((1,4))
#edge2.add((1,2))
#edge2.add((1,3))
#edge2.add((2,3))
#print edge1
#print edge2
#print edge1-edge2
#print edge2-edge1